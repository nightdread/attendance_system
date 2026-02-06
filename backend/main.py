from fastapi import FastAPI, HTTPException, Depends, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse, Response
import json
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import uvicorn
import os
import sys
import time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from pathlib import Path
# Add project root to path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

from config import (
    BOT_USERNAME, WEB_USERNAME, WEB_PASSWORD,
    API_HOST, API_PORT, QR_UPDATE_INTERVAL, SECRET_KEY, SESSION_SECRET_KEY, API_KEY, DB_PATH
)
from config.config import ADMIN_IP_WHITELIST
from database import Database
from auth.jwt_handler import JWTHandler
from backend.schemas import (
    TokenResponse, ErrorResponse, UserResponse, UserUpdateRequest,
    HealthCheckResponse, MetricsResponse, AnalyticsSummaryResponse,
    DailyStatsResponse, EmployeeStatsResponse
)
from utils.logger import (
    log_error, log_request, log_auth_event,
    log_failed_login, log_successful_login, log_role_change,
    log_suspicious_activity, log_rate_limit_exceeded, log_csrf_failure,
    log_unauthorized_access, log_data_export
)
from utils.cache import cache
from utils.validators import (
    validate_username, validate_password, validate_fio, validate_role,
    validate_department, validate_position, sanitize_string
)
from utils.rate_limit import rate_limit
from utils.csrf import set_csrf_token, get_csrf_token, require_csrf_token

app = FastAPI(
    title="Attendance System API",
    description="""
    API для системы учета рабочего времени.
    
    ## Аутентификация
    
    API поддерживает два способа аутентификации:
    1. **API Key** - через заголовок `X-API-Key` (если настроен API_KEY)
    2. **JWT Token** - через сессию или заголовок `Authorization: Bearer <token>`
    
    ## Rate Limiting
    
    Endpoints имеют ограничения по частоте запросов:
    - `/api/active_token`: 10 запросов/минуту
    - `/api/user/*`: 20 запросов/минуту (GET), 10 запросов/минуту (PUT)
    - `/api/analytics/*`: 30 запросов/минуту
    - `/login`: 5 попыток/5 минут
    
    При превышении лимита возвращается HTTP 429.
    
    ## CSRF Protection
    
    Все модифицирующие операции (POST, PUT, DELETE) требуют CSRF токен:
    - Для форм: поле `csrf_token`
    - Для JSON: заголовок `X-CSRF-Token`
    
    ## Версионирование
    
    Текущая версия API: **v1**
    
    ## Ошибки
    
    API возвращает стандартные HTTP коды:
    - `200` - Успех
    - `400` - Неверный запрос (валидация)
    - `401` - Не авторизован
    - `403` - Доступ запрещен (недостаточно прав или CSRF)
    - `404` - Не найдено
    - `429` - Превышен rate limit
    - `500` - Внутренняя ошибка сервера
    - `503` - Сервис недоступен (health check)
    """,
    version="1.0.0",
    contact={
        "name": "Attendance System Support",
    },
    license_info={
        "name": "Proprietary",
    },
    tags_metadata=[
        {
            "name": "authentication",
            "description": "Аутентификация и авторизация",
        },
        {
            "name": "tokens",
            "description": "Управление токенами для отметки посещаемости",
        },
        {
            "name": "users",
            "description": "Управление пользователями (требует admin/manager роль)",
        },
        {
            "name": "analytics",
            "description": "Аналитика и статистика (требует admin/manager/hr роль)",
        },
        {
            "name": "health",
            "description": "Проверка здоровья системы и метрики",
        },
    ]
)

# Add session middleware (use dedicated session secret, defaulting to SECRET_KEY)
# max_age=31536000 = 1 year in seconds for long-lived sessions (terminal)
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET_KEY, max_age=31536000)

# Login rate limiter (Redis + fallback memory)
MAX_LOGIN_ATTEMPTS = 5
LOGIN_WINDOW_SEC = 300  # 5 minutes
LOGIN_BLOCK_SEC = 900   # block 15 minutes
LOGIN_ATTEMPTS = defaultdict(list)


def check_ip_whitelist(request: Request) -> bool:
    """Проверить IP адрес против whitelist администраторов"""
    if not ADMIN_IP_WHITELIST:
        return True  # Если whitelist не настроен, разрешаем всем
    
    client_ip = request.client.host if request.client else None
    if not client_ip:
        return False
    
    # Проверяем прямой IP и X-Forwarded-For заголовок
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        # Берем первый IP из цепочки прокси
        client_ip = forwarded_for.split(",")[0].strip()
    
    return client_ip in ADMIN_IP_WHITELIST

def authorize_request(
    request: Request,
    require_roles: list[str] | None = None,
    allow_terminal_session: bool = False,
):
    """Authorize API requests.

    - If API_KEY is set, require matching X-API-Key header.
    - Else, allow authenticated session (access_token in session or bearer header).
    - Optionally allow terminal session flag (set on public terminal page) for read-only endpoints.
    - If ADMIN_IP_WHITELIST is set and require_roles includes "admin", check IP whitelist.
    """
    # API key check (if configured and header provided)
    if API_KEY:
        header_key = request.headers.get("X-API-Key")
        if header_key == API_KEY:
            return {"role": "api"}

    # Allow terminal session flag (for public terminal QR updates) if enabled
    if allow_terminal_session and request.session.get("terminal_allowed"):
        return {"role": "terminal"}

    # Session or bearer token
    token = request.session.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header.split(" ", 1)[1]

    if not token:
        client_ip = request.client.host if request.client else "unknown"
        log_unauthorized_access(str(request.url.path), ip_address=client_ip, reason="No token provided")
        raise HTTPException(status_code=401, detail="Unauthorized")

    payload = JWTHandler.verify_token(token)
    if not payload:
        client_ip = request.client.host if request.client else "unknown"
        log_unauthorized_access(str(request.url.path), ip_address=client_ip, reason="Invalid token")
        raise HTTPException(status_code=401, detail="Unauthorized")

    role = payload.get("role", "user")
    username = payload.get("sub")
    
    # Проверка IP whitelist для администраторов
    if require_roles and "admin" in require_roles and role == "admin":
        if not check_ip_whitelist(request):
            client_ip = request.client.host if request.client else "unknown"
            log_unauthorized_access(str(request.url.path), user=username, ip_address=client_ip, reason="IP not in admin whitelist")
            raise HTTPException(status_code=403, detail="Forbidden: IP address not allowed")
    
    if require_roles and role not in require_roles:
        client_ip = request.client.host if request.client else "unknown"
        log_unauthorized_access(str(request.url.path), user=username, ip_address=client_ip, reason=f"Insufficient permissions: role '{role}' not in {require_roles}")
        raise HTTPException(status_code=403, detail="Forbidden")

    return payload

# Resolve paths that work both locally and inside Docker
BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"

# Mount static files
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Templates
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Add custom Jinja2 filters
def format_hours_filter(hours: float) -> str:
    """Jinja2 filter to format hours as HH:MM"""
    return format_hours_to_hhmm_util(hours)

templates.env.filters["format_hours"] = format_hours_filter

# Database instance
db = Database(str(DB_PATH))

# Dependency to get database
def get_db():
    return db

def build_terminal_context(request: Request, db: Database) -> dict:
    """Prepare context for the public terminal page"""
    token_data = db.get_active_token()
    token = token_data['token'] if token_data else db.create_token()
    url = f"https://t.me/{BOT_USERNAME}?start={token}"

    # Проверяем авторизацию для отображения админских ссылок (если пользователь уже логинился)
    user_info = None
    try:
        jwt_token = request.session.get("access_token")
        if jwt_token:
            payload = JWTHandler.verify_token(jwt_token)
            username = payload.get("sub")
            if username:
                user = db.get_web_user_by_username(username)
                if user:
                    user_info = {
                        "username": username,
                        "role": user.get("role", "user"),
                        "is_admin": user.get("role") in ["admin", "manager", "hr"]
                    }
    except Exception:
        # если токен невалиден, просто показываем публичную страницу
        pass

    return {
            "request": request,
            "token": token,
            "url": url,
            "bot_url": url,
            "update_interval": QR_UPDATE_INTERVAL * 1000,
            "user_info": user_info
        }


@app.get(
    "/api/active_token",
    response_model=TokenResponse,
    responses={
        200: {"description": "Success", "model": TokenResponse},
        401: {"description": "Unauthorized - требуется аутентификация", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded - превышен лимит запросов (10/мин)", "model": ErrorResponse}
    },
    tags=["tokens"],
    summary="Получить активный токен для отметки",
    description="""
    Возвращает текущий активный токен для отметки посещаемости через Telegram бота.
    
    **Аутентификация:** Требуется (API key или JWT token)
    
    **Rate Limit:** 10 запросов в минуту
    
    **Пример использования:**
    ```bash
    curl -H "Authorization: Bearer <token>" https://api.example.com/api/active_token
    ```
    """
)
async def get_active_token(request: Request, db: Database = Depends(get_db)):
    """Get active token for attendance"""
    # Rate limiting
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="api_token")
    # Авторизация: API key или авторизованная сессия
    # Терминал теперь требует авторизацию, поэтому убираем allow_terminal_session
    authorize_request(request, allow_terminal_session=False)

    token_data = db.get_active_token()
    token = token_data['token'] if token_data else db.create_token()
    url = f"https://t.me/{BOT_USERNAME}?start={token}"
    return {"token": token, "url": url, "bot_url": url}

@app.get("/api/token")
async def get_token_for_device(request: Request, db: Database = Depends(get_db)):
    """Get active token for microcontroller/device (simplified endpoint)
    
    This endpoint is designed for microcontrollers (ESP32, Arduino, etc.)
    that need to get the current token and generate QR code locally.
    
    Authentication: Optional API key via X-API-Key header (if API_KEY is configured)
    If no API_KEY is set, this endpoint is publicly accessible (read-only).
    
    Returns:
        {
            "token": "current_token_string",
            "url": "https://t.me/bot_username?start=token",
            "bot_username": "bot_username"
        }
    """
    # Optional API key check (if configured)
    if API_KEY:
        header_key = request.headers.get("X-API-Key")
        if header_key != API_KEY:
            raise HTTPException(status_code=401, detail="Invalid API key")
    
    token_data = db.get_active_token()
    token = token_data['token'] if token_data else db.create_token()
    url = f"https://t.me/{BOT_USERNAME}?start={token}"
    
    # Include token creation timestamp for change detection
    if token_data:
        created_at = token_data.get('created_at', '')
    else:
        # If we just created a token, get it again to include timestamp
        token_data = db.get_active_token()
        created_at = token_data.get('created_at', '') if token_data else ''
    
    return {
        "token": token,
        "url": url,
        "bot_username": BOT_USERNAME,
        "created_at": created_at  # ISO format timestamp for change detection
    }

# Web terminal routes
@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Database = Depends(get_db)):
    """Главная: терминал с QR-кодом (требует авторизацию с ролью terminal)"""
    # Проверяем авторизацию
    if not request.session.get("authenticated"):
        return RedirectResponse(url="/login?next=/", status_code=302)
    
    # Проверяем роль - только terminal может видеть QR-код
    user_role = request.session.get("user_role")
    if user_role != "terminal":
        # Если не terminal, редиректим на админку или страницу доступа
        if user_role in ["admin", "manager", "hr"]:
            return RedirectResponse(url="/admin", status_code=302)
        else:
            return RedirectResponse(url="/me", status_code=302)
    
    context = build_terminal_context(request, db)
    return templates.TemplateResponse("terminal.html", context)

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Страница входа с поддержкой редиректа после логина"""
    next_url = request.query_params.get("next", "/terminal")
    # Генерируем CSRF токен для формы логина
    csrf_token = set_csrf_token(request)
    return templates.TemplateResponse("login.html", {
        "request": request,
        "next_url": next_url,
        "csrf_token": csrf_token
    })

@app.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Database = Depends(get_db),
    next_url: str = Form("/terminal"),
    csrf_token: str = Form(None)
):
    # Проверка CSRF токена (передаем токен из формы)
    await require_csrf_token(request, form_token=csrf_token)
    
    # Валидация входных данных
    username = sanitize_string(username, max_length=50)
    is_valid, error_msg = validate_username(username)
    if not is_valid:
        csrf_token = get_csrf_token(request) or set_csrf_token(request)
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": error_msg, "next_url": next_url, "csrf_token": csrf_token}
        )
    
    is_valid, error_msg = validate_password(password, min_length=6, require_complexity=False)
    if not is_valid:
        csrf_token = get_csrf_token(request) or set_csrf_token(request)
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": error_msg, "next_url": next_url, "csrf_token": csrf_token}
        )
    
    # Rate limit by client IP
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()

    # Redis-based limiter (fallback to memory)
    try:
        if cache.redis_client:
            key_block = f"login:block:{client_ip}"
            key_counter = f"login:count:{client_ip}"

            if cache.redis_client.exists(key_block):
                return templates.TemplateResponse(
                    "login.html",
                    {
                        "request": request,
                        "error": "Слишком много попыток. Попробуйте позже.",
                        "next_url": next_url
                    }
                )

            count = cache.redis_client.incr(key_counter)
            if count == 1:
                cache.redis_client.expire(key_counter, LOGIN_WINDOW_SEC)
            if count > MAX_LOGIN_ATTEMPTS:
                cache.redis_client.set(key_block, 1, ex=LOGIN_BLOCK_SEC)
                log_rate_limit_exceeded("/login", client_ip, attempts=count)
                return templates.TemplateResponse(
                    "login.html",
                    {
                        "request": request,
                        "error": "Слишком много попыток. Попробуйте позже.",
                        "next_url": next_url
                    }
                )
        else:
            # Use get() to avoid KeyError if key doesn't exist
            attempts = LOGIN_ATTEMPTS.get(client_ip, [])
            LOGIN_ATTEMPTS[client_ip] = [t for t in attempts if now - t < LOGIN_WINDOW_SEC]
            if len(LOGIN_ATTEMPTS.get(client_ip, [])) >= MAX_LOGIN_ATTEMPTS:
                log_rate_limit_exceeded("/login", client_ip, attempts=len(LOGIN_ATTEMPTS.get(client_ip, [])))
                return templates.TemplateResponse(
                    "login.html",
                    {
                        "request": request,
                        "error": "Слишком много попыток. Попробуйте позже.",
                        "next_url": next_url
                    }
                )
    except Exception as e:
        # Log rate limiting errors but don't block login
        log_error(e, "Rate limiting")

    # Попытка логина через API
    try:
        # Используем authenticate_web_user, который обновляет last_login
        user = db.authenticate_web_user(username, password)
        if user:
            # Сброс счетчика попыток при успехе
            try:
                if cache.redis_client:
                    cache.redis_client.delete(f"login:count:{client_ip}")
                    cache.redis_client.delete(f"login:block:{client_ip}")
                else:
                    LOGIN_ATTEMPTS.pop(client_ip, None)
            except Exception as e:
                log_error(e, "Reset login attempts counter")
            # Генерируем JWT токен
            # Для роли terminal делаем очень долгую сессию (1 год) для работы 24/7
            user_role = user.get("role", "user")
            if user_role == "terminal":
                # Для терминала - сессия на 1 год (525600 минут)
                expires_delta = timedelta(days=365)
            else:
                # Для остальных ролей - стандартное время (30 минут)
                expires_delta = None
            
            access_token = JWTHandler.create_access_token(
                data={"sub": username, "role": user_role},
                expires_delta=expires_delta
            )
            # Сохраняем токен в сессии
            request.session["access_token"] = access_token
            request.session["authenticated"] = True
            request.session["user_role"] = user.get("role", "user")
            # Генерируем новый CSRF токен после успешного логина
            set_csrf_token(request)
            # Логируем успешный вход
            log_successful_login(username, client_ip, role=user_role)

            # Редирект на запрошенную страницу или по умолчанию
            # Админы могут попасть в админку, но по умолчанию все идут на терминал
            redirect_to = next_url if next_url and next_url.startswith("/") else "/terminal"
            return RedirectResponse(url=redirect_to, status_code=302)
        else:
            # Initialize list if key doesn't exist (defaultdict handles this, but being explicit)
            if client_ip not in LOGIN_ATTEMPTS:
                LOGIN_ATTEMPTS[client_ip] = []
            LOGIN_ATTEMPTS[client_ip].append(now)
            # Логируем неудачную попытку входа
            log_failed_login(username, client_ip, reason="Invalid credentials")
            return templates.TemplateResponse(
                "login.html",
                {"request": request, "error": "Invalid credentials", "next_url": next_url}
            )
    except Exception as e:
        log_error(e, "Login")
        # Initialize list if key doesn't exist (defaultdict handles this, but being explicit)
        if client_ip not in LOGIN_ATTEMPTS:
            LOGIN_ATTEMPTS[client_ip] = []
        LOGIN_ATTEMPTS[client_ip].append(now)
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Login failed", "next_url": next_url}
        )

@app.get("/terminal", response_class=HTMLResponse)
async def terminal_page(request: Request, db: Database = Depends(get_db)):
    """Терминал с QR-кодом (требует авторизацию с ролью terminal)"""
    # Проверяем авторизацию
    if not request.session.get("authenticated"):
        return RedirectResponse(url="/login?next=/terminal", status_code=302)
    
    # Проверяем роль - только terminal может видеть QR-код
    user_role = request.session.get("user_role")
    if user_role != "terminal":
        # Если не terminal, редиректим на админку или страницу доступа
        if user_role in ["admin", "manager", "hr"]:
            return RedirectResponse(url="/admin", status_code=302)
        else:
            return RedirectResponse(url="/me", status_code=302)
    
    context = build_terminal_context(request, db)
    return templates.TemplateResponse("terminal.html", context)

@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=302)

# Admin routes for reporting
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request, db: Database = Depends(get_db)):
    if not request.session.get("authenticated"):
        return RedirectResponse(url="/login", status_code=302)

    user_role = request.session.get("user_role", "user")

    # Get currently present users
    present_users = db.get_currently_present()

    initial_creds = None
    if user_role in ["admin", "manager", "hr"]:
        initial_creds = db.consume_initial_credentials()

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "present_users": present_users,
            "initial_creds": initial_creds
        }
    )

@app.get("/admin/user/{user_id}", response_class=HTMLResponse)
async def user_history(request: Request, user_id: int, db: Database = Depends(get_db)):
    if not request.session.get("authenticated"):
        return RedirectResponse(url="/login", status_code=302)

    # Get user info
    user = db.get_person_by_tg_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Get user events
    events = db.get_user_events(user_id, limit=100)

    return templates.TemplateResponse(
        "user_history.html",
        {
            "request": request,
            "user": user,
            "events": events
        }
    )

@app.get("/me", response_class=HTMLResponse)
async def self_dashboard(request: Request, db: Database = Depends(get_db)):
    """Personal dashboard for authenticated users (role user and above)"""
    token = request.session.get("access_token")
    if not token:
        return RedirectResponse(url="/login", status_code=302)

    try:
        payload = JWTHandler.verify_token(token)
        username = payload.get("sub")
        role = payload.get("role", "user")

        # derive tg_user_id from username pattern user<tgid>
        tg_user_id = None
        if username and username.startswith("user"):
            suffix = username[4:]
            if suffix.isdigit():
                tg_user_id = int(suffix)

        if tg_user_id is None:
            return RedirectResponse(url="/login", status_code=302)

        person = db.get_person_by_tg_id(tg_user_id)
        stats = db.get_employee_stats_by_tg(tg_user_id)

        if not person or not stats:
            return templates.TemplateResponse(
                "self_dashboard.html",
                {
                    "request": request,
                    "error": "Профиль не найден. Отсканируйте QR через бота, чтобы зарегистрироваться."
                }
            )

        return templates.TemplateResponse(
            "self_dashboard.html",
            {
                "request": request,
                "person": person,
                "stats": stats,
                "role": role
            }
        )
    except Exception as e:
        log_error(e, "Self dashboard")
        request.session.clear()
        return RedirectResponse(url="/login", status_code=302)

# Analytics and user management pages
@app.get("/analytics", response_class=HTMLResponse)
async def analytics_dashboard(request: Request):
    """Analytics dashboard page"""
    token = request.session.get("access_token")
    if not token:
        return RedirectResponse(url="/login", status_code=302)

    try:
        payload = JWTHandler.verify_token(token)
        user_role = payload.get("role", "user")

        # Проверяем права доступа
        if user_role not in ["admin", "manager", "hr"]:
            return RedirectResponse(url="/terminal", status_code=302)

        # Get analytics data
        analytics_summary = db.get_analytics_summary()
        daily_visits = db.get_daily_visits_chart()
        hourly_distribution = db.get_hourly_distribution()
        top_workers = db.get_top_workers()
        department_stats = db.get_department_stats()

        # Get employee list for selection
        employee_list = db.get_employee_list()

        return templates.TemplateResponse(
            "analytics.html",
            {
                "request": request,
                "analytics_summary": analytics_summary,
                "daily_visits": daily_visits,
                "hourly_distribution": hourly_distribution,
                "top_workers": top_workers,
                "department_stats": department_stats,
                "employee_list": employee_list
            }
        )
    except Exception as e:
        log_error(e, "Analytics page token validation")
        request.session.clear()
        return RedirectResponse(url="/login", status_code=302)

@app.get("/users", response_class=HTMLResponse)
async def user_management(request: Request, db: Database = Depends(get_db)):
    """User management page"""
    token = request.session.get("access_token")
    if not token:
        return RedirectResponse(url="/login", status_code=302)

    # Генерируем CSRF токен для форм на странице
    csrf_token = set_csrf_token(request)

    try:
        payload = JWTHandler.verify_token(token)
        user_role = payload.get("role", "user")

        # Проверяем права доступа (только админ и менеджер)
        if user_role not in ["admin", "manager"]:
            client_ip = request.client.host if request.client else "unknown"
            username = payload.get("sub")
            log_unauthorized_access("/users", user=username, ip_address=client_ip, reason=f"Role '{user_role}' not allowed")
            return RedirectResponse(url="/terminal", status_code=302)

        # Получаем данные для шаблона
        users = db.get_all_web_users()
        roles = db.get_all_roles()

        return templates.TemplateResponse(
            "user_management.html",
            {
                "request": request,
                "users": users,
                "roles": roles,
                "current_user_role": user_role,
                "csrf_token": csrf_token
            }
        )
    except Exception as e:
        log_error(e, "User management page")
        request.session.clear()
        return RedirectResponse(url="/login", status_code=302)

# API endpoints for user management
@app.get(
    "/api/user/{user_id}",
    response_model=UserResponse,
    responses={
        200: {"description": "Success", "model": UserResponse},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden - требуется роль admin или manager", "model": ErrorResponse},
        404: {"description": "User not found", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["users"],
    summary="Получить пользователя по ID",
    description="""
    Возвращает информацию о пользователе по его ID.
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin или manager
    - Rate Limit: 20 запросов/минуту
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" https://api.example.com/api/user/1
    ```
    """
)
async def get_user(request: Request, user_id: int, db: Database = Depends(get_db)):
    """Get user by ID"""
    rate_limit(request, max_requests=20, window_seconds=60, key_prefix="user_mgmt")
    authorize_request(request, require_roles=["admin", "manager"])
    
    user = db.get_web_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Remove password hash from response
    user.pop('password_hash', None)
    return user

@app.put(
    "/api/user/{user_id}",
    response_model=UserResponse,
    responses={
        200: {"description": "Success", "model": UserResponse},
        400: {"description": "Bad Request - ошибка валидации данных", "model": ErrorResponse},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden - требуется роль admin/manager или неверный CSRF токен", "model": ErrorResponse},
        404: {"description": "User not found", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded - превышен лимит (10/мин)", "model": ErrorResponse}
    },
    tags=["users"],
    summary="Обновить пользователя",
    description="""
    Обновляет информацию о пользователе.
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin или manager
    - CSRF токен: обязателен (заголовок `X-CSRF-Token`)
    - Rate Limit: 10 запросов/минуту
    
    **Валидация:**
    - `full_name`: 3-200 символов, только буквы, пробелы, дефисы
    - `role`: один из: user, admin, manager, hr, terminal
    - `password`: минимум 8 символов, должен содержать буквы и цифры
    - `department`, `position`: максимум 100 символов
    
    **Пример запроса:**
    ```json
    {
        "full_name": "Иванов Иван Иванович",
        "role": "user",
        "department": "IT",
        "is_active": true
    }
    ```
    
    **Пример curl:**
    ```bash
    curl -X PUT https://api.example.com/api/user/1 \\
      -H "Authorization: Bearer <token>" \\
      -H "X-CSRF-Token: <csrf_token>" \\
      -H "Content-Type: application/json" \\
      -d '{"full_name": "New Name", "role": "user"}'
    ```
    """
)
async def update_user(
    request: Request,
    user_id: int,
    db: Database = Depends(get_db)
):
    """Update user"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="user_mgmt")
    
    # Parse JSON body first to get CSRF token
    form_token = None
    try:
        body = await request.json()
        # For JSON requests, CSRF token should be in header
    except:
        # Fallback to form data
        form = await request.form()
        body = dict(form)
        # Get CSRF token from form if present
        form_token = body.get("csrf_token")
    
    # Проверка CSRF токена (передаем токен из формы, если есть)
    await require_csrf_token(request, form_token=form_token)
    # Authorize and get payload once
    payload = authorize_request(request, require_roles=["admin", "manager"])
    
    # Get current user ID for audit
    current_username = payload.get("sub")
    current_user = db.get_web_user_by_username(current_username) if current_username else None
    updated_by = current_user.get("id") if current_user else None
    
    # Check if user exists
    user = db.get_web_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        # Convert is_active string to boolean
        if 'is_active' in body:
            body['is_active'] = body['is_active'].lower() in ('true', '1', 'yes', 'on')
    
    # Extract fields (only update provided fields)
    full_name = body.get('full_name')
    role = body.get('role')
    department = body.get('department')
    position = body.get('position')
    is_active = body.get('is_active')
    password = body.get('password')
    
    # Валидация полей
    if full_name is not None:
        full_name = sanitize_string(str(full_name), max_length=200)
        is_valid, error_msg = validate_fio(full_name)
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Ошибка валидации ФИО: {error_msg}")
    
    if role is not None:
        is_valid, error_msg = validate_role(str(role))
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Ошибка валидации роли: {error_msg}")
        # Логируем изменение роли
        old_role = user.get("role")
        if old_role != role:
            client_ip = request.client.host if request.client else "unknown"
            log_role_change(current_username or "unknown", user.get("username", f"user_{user_id}"), old_role, role, client_ip)
    
    if department is not None:
        department = sanitize_string(str(department), max_length=100)
        is_valid, error_msg = validate_department(department)
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Ошибка валидации отдела: {error_msg}")
    
    if position is not None:
        position = sanitize_string(str(position), max_length=100)
        is_valid, error_msg = validate_position(position)
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Ошибка валидации должности: {error_msg}")
    
    # Remove empty password
    if password and password.strip() == '':
        password = None
    elif password:
        is_valid, error_msg = validate_password(password, min_length=8, require_complexity=True)
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Ошибка валидации пароля: {error_msg}")
    
    # Update user
    success = db.update_web_user(
        user_id=user_id,
        full_name=full_name,
        role=role,
        department=department,
        position=position,
        is_active=is_active,
        password=password,
        updated_by=updated_by
    )
    
    if not success:
        raise HTTPException(status_code=400, detail="Failed to update user")
    
    # Return updated user
    updated_user = db.get_web_user_by_id(user_id)
    updated_user.pop('password_hash', None)
    return updated_user

@app.get(
    "/api/employee/{employee_id}",
    responses={
        200: {"description": "Success"},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        404: {"description": "Employee not found", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Статистика по сотруднику",
    description="""
    Возвращает детальную статистику для конкретного сотрудника.
    
    **Параметры:**
    - `employee_id`: ID сотрудника (Telegram user ID)
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" \\
      https://api.example.com/api/employee/123456
    ```
    """
)
async def get_employee_stats(employee_id: int, db: Database = Depends(get_db)):
    """Get detailed statistics for a specific employee"""
    stats = db.get_employee_detailed_stats(employee_id)
    if not stats:
        raise HTTPException(status_code=404, detail="Employee not found")

    return stats

@app.get("/api/employees/date/{date}")
async def get_employees_by_date(request: Request, date: str, db: Database = Depends(get_db)):
    """Get list of employees who visited on a specific date (YYYY-MM-DD)"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    try:
        # Validate date format
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    employees = db.get_employees_by_date(date)
    return {
        "date": date,
        "employees": employees,
        "total": len(employees)
    }

@app.get(
    "/api/analytics/daily/{date}",
    responses={
        200: {"description": "Success"},
        400: {"description": "Bad Request - неверный формат даты", "model": ErrorResponse},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Дневная аналитика",
    description="""
    Возвращает статистику посещаемости за указанную дату.
    
    **Параметры:**
    - `date`: Дата в формате YYYY-MM-DD
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    - Rate Limit: 30 запросов/минуту
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" \\
      https://api.example.com/api/analytics/daily/2024-01-15
    ```
    """
)
async def analytics_daily(request: Request, date: str, db: Database = Depends(get_db)):
    """Daily analytics by date (YYYY-MM-DD)"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    stats = db.get_daily_stats(date)
    return {"date": date, **stats}

@app.get(
    "/api/analytics/weekly",
    responses={
        200: {"description": "Success"},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Недельная аналитика",
    description="""
    Возвращает статистику посещаемости за последние 7 дней.
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    - Rate Limit: 30 запросов/минуту
    """
)
async def analytics_weekly(request: Request, db: Database = Depends(get_db)):
    """Weekly analytics for last 7 days"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    today = datetime.utcnow().date()
    start_date = (today - timedelta(days=6)).isoformat()
    end_date = today.isoformat()
    data = db.get_weekly_stats(start_date, end_date)
    return {
        "period": {"start": start_date, "end": end_date},
        "data": data,
        "daily_stats": data
    }

@app.get(
    "/api/analytics/locations",
    responses={
        200: {"description": "Success"},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Аналитика по локациям",
    description="""
    Возвращает статистику по локациям (в текущей реализации все события глобальные).
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    - Rate Limit: 30 запросов/минуту
    """
)
async def analytics_locations(request: Request, db: Database = Depends(get_db)):
    """Analytics by locations (global in current implementation)"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    data = db.get_location_stats()
    return {"locations": data}

@app.get(
    "/api/analytics/users",
    responses={
        200: {"description": "Success"},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Самые активные пользователи",
    description="""
    Возвращает список самых активных пользователей.
    
    **Параметры:**
    - `limit`: Количество пользователей (по умолчанию 10)
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    - Rate Limit: 30 запросов/минуту
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" \\
      "https://api.example.com/api/analytics/users?limit=20"
    ```
    """
)
async def analytics_users(request: Request, limit: int = 10, db: Database = Depends(get_db)):
    """Most active users"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    data = db.get_user_stats(limit)
    return {"users": data}

@app.get(
    "/api/analytics/hourly/{date}",
    responses={
        200: {"description": "Success"},
        400: {"description": "Bad Request - неверный формат даты", "model": ErrorResponse},
        401: {"description": "Unauthorized", "model": ErrorResponse},
        403: {"description": "Forbidden", "model": ErrorResponse},
        429: {"description": "Rate limit exceeded", "model": ErrorResponse}
    },
    tags=["analytics"],
    summary="Почасовая аналитика",
    description="""
    Возвращает распределение посещаемости по часам за указанную дату.
    
    **Параметры:**
    - `date`: Дата в формате YYYY-MM-DD
    
    **Требования:**
    - Аутентификация: обязательна
    - Роль: admin, manager или hr
    - Rate Limit: 30 запросов/минуту
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" \\
      https://api.example.com/api/analytics/hourly/2024-01-15
    ```
    """
)
async def analytics_hourly(request: Request, date: str, db: Database = Depends(get_db)):
    """Hourly analytics for a date (YYYY-MM-DD)"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    data = db.get_hourly_stats(date)
    return {"date": date, "hourly": data, "hourly_stats": data}

@app.get(
    "/api/health",
    response_model=HealthCheckResponse,
    responses={
        200: {"description": "System is healthy", "model": HealthCheckResponse},
        503: {"description": "System is degraded - некоторые сервисы недоступны", "model": HealthCheckResponse}
    },
    tags=["health"],
    summary="Проверка здоровья системы",
    description="""
    Проверяет состояние системы и всех зависимостей.
    
    **Проверяет:**
    - База данных (подключение, размер)
    - Redis (если включен)
    - Системные метрики (CPU, память, диск) - опционально
    
    **Статусы:**
    - `healthy` (200) - все сервисы работают
    - `degraded` (503) - некоторые сервисы недоступны
    
    **Использование:**
    - Docker health checks
    - Мониторинг систем
    - Load balancer health checks
    """
)
async def health_check(db: Database = Depends(get_db)):
    """Enhanced health check endpoint with detailed system information"""
    from utils.metrics import get_system_metrics, get_redis_metrics, get_database_metrics
    
    health_status = {
        "status": "healthy",
        "version": "1.0.0",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "checks": {}
    }
    
    overall_healthy = True
    
    # Database check
    try:
        db_metrics = get_database_metrics(db)
        health_status["checks"]["database"] = db_metrics
        if db_metrics.get("status") != "healthy":
            overall_healthy = False
    except Exception as e:
        health_status["checks"]["database"] = {"status": "error", "error": str(e)}
        overall_healthy = False
    
    # Redis check
    try:
        redis_metrics = get_redis_metrics()
        health_status["checks"]["redis"] = redis_metrics
        from config.config import REDIS_ENABLED
        if REDIS_ENABLED and cache.redis_client and not redis_metrics.get("connected"):
            overall_healthy = False
    except Exception as e:
        health_status["checks"]["redis"] = {"status": "error", "error": str(e)}
        from config.config import REDIS_ENABLED
        if REDIS_ENABLED and cache.redis_client:
            overall_healthy = False
    
    # System metrics (optional, don't fail if unavailable)
    try:
        system_metrics = get_system_metrics()
        if "error" not in system_metrics:
            health_status["system"] = system_metrics
    except:
        pass  # System metrics are optional
    
    if not overall_healthy:
        health_status["status"] = "degraded"
    
    status_code = 200 if overall_healthy else 503
    return Response(
        content=json.dumps(health_status, indent=2),
        media_type="application/json",
        status_code=status_code
    )

@app.get("/api/analytics/compare")
async def analytics_compare(
    request: Request,
    period1_start: str,
    period1_end: str,
    period2_start: str,
    period2_end: str,
    db: Database = Depends(get_db)
):
    """Сравнить два периода по метрикам"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    try:
        datetime.strptime(period1_start, "%Y-%m-%d")
        datetime.strptime(period1_end, "%Y-%m-%d")
        datetime.strptime(period2_start, "%Y-%m-%d")
        datetime.strptime(period2_end, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    return db.compare_periods(period1_start, period1_end, period2_start, period2_end)

@app.get("/api/analytics/late-arrivals")
async def analytics_late_arrivals(
    request: Request,
    start_date: str,
    end_date: str,
    late_threshold_hours: int = 9,
    db: Database = Depends(get_db)
):
    """Получить статистику опозданий"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    return db.get_late_arrivals_stats(start_date, end_date, late_threshold_hours)

@app.get("/api/analytics/overtime")
async def analytics_overtime(
    request: Request,
    start_date: str,
    end_date: str,
    standard_hours_per_day: float = 8.0,
    db: Database = Depends(get_db)
):
    """Получить отчет по переработкам"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    return db.get_overtime_report(start_date, end_date, standard_hours_per_day)

@app.get("/api/analytics/weekly-distribution")
async def analytics_weekly_distribution(
    request: Request,
    start_date: str,
    end_date: str,
    db: Database = Depends(get_db)
):
    """Получить распределение рабочего времени по дням недели"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    return db.get_weekly_distribution(start_date, end_date)

@app.get("/api/analytics/calendar/{year}/{month}")
async def analytics_calendar(
    request: Request,
    year: int,
    month: int,
    db: Database = Depends(get_db)
):
    """Получить данные календаря для конкретного месяца"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_analytics")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Invalid month. Must be 1-12")
    
    # Вычисляем диапазон дат для месяца
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    start_datetime = f"{start_date.isoformat()}T00:00:00"
    end_datetime = f"{end_date.isoformat()}T23:59:59"
    
    # Получаем данные о посещениях и часах работы за месяц
    with db.get_connection() as conn:
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT
                date(ts) as visit_date,
                COUNT(DISTINCT user_id) as visits
            FROM events
            WHERE ts >= ? AND ts <= ? AND action = 'in'
            GROUP BY date(ts)
            ORDER BY date(ts)
        """, (start_datetime, end_datetime))
        
        visits_data = {row[0]: row[1] for row in cursor.fetchall()}
        
        # Получаем все события за период для правильного расчета часов работы
        cursor.execute("""
            SELECT
                user_id,
                date(ts) as work_date,
                ts,
                action
            FROM events
            WHERE ts >= ? AND ts <= ? AND action IN ('in', 'out')
            ORDER BY user_id, ts
        """, (start_datetime, end_datetime))
        
        # Правильно считаем часы работы с учетом всех интервалов
        hours_data = {}
        user_daily_events = {}
        
        for row in cursor.fetchall():
            user_id = row[0]
            date_str = row[1]
            ts_str = row[2]
            action = row[3]
            
            key = (user_id, date_str)
            if key not in user_daily_events:
                user_daily_events[key] = []
            
            user_daily_events[key].append({
                'ts': ts_str,
                'action': action
            })
        
        # Считаем часы работы для каждого пользователя и дня
        for (user_id, date_str), events in user_daily_events.items():
            total_seconds = 0
            checkin_time = None
            
            # Сортируем события по времени
            sorted_events = sorted(events, key=lambda x: x['ts'])
            
            for event in sorted_events:
                event_time = datetime.fromisoformat(event['ts'].replace('Z', '+00:00'))
                
                if event['action'] == 'in':
                    checkin_time = event_time
                elif event['action'] == 'out' and checkin_time:
                    checkout_time = event_time
                    work_seconds = (checkout_time - checkin_time).total_seconds()
                    if 0 < work_seconds < 86400:  # Валидация: от 0 до 24 часов
                        total_seconds += work_seconds
                    checkin_time = None
            
            if total_seconds > 0:
                hours = total_seconds / 3600.0
                if date_str not in hours_data:
                    hours_data[date_str] = 0
                hours_data[date_str] += hours
        
        # Проверяем выходные и праздничные дни через производственный календарь
        is_holiday_data = {}
        try:
            from utils.production_calendar import is_working_day
            current_check = start_date
            while current_check <= end_date:
                date_str = current_check.isoformat()
                is_holiday_data[date_str] = not is_working_day(current_check)
                current_check += timedelta(days=1)
        except Exception as e:
            # Если производственный календарь недоступен, используем стандартную логику (суббота/воскресенье)
            current_check = start_date
            while current_check <= end_date:
                date_str = current_check.isoformat()
                weekday = current_check.weekday()  # 0 = Monday, 6 = Sunday
                is_holiday_data[date_str] = (weekday >= 5)  # Saturday (5) or Sunday (6)
                current_check += timedelta(days=1)
        
        # Объединяем данные
        result = []
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.isoformat()
            result.append({
                'date': date_str,
                'visits': visits_data.get(date_str, 0),
                'hours': round(hours_data.get(date_str, 0), 1),
                'is_holiday': is_holiday_data.get(date_str, False)
            })
            current_date += timedelta(days=1)
        
        return {'days': result}

@app.get("/api/audit-log")
async def get_audit_log(
    request: Request,
    limit: int = 100,
    offset: int = 0,
    action_type: str = None,
    user_id: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Database = Depends(get_db)
):
    """Получить записи журнала аудита"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_audit")
    authorize_request(request, require_roles=["admin"])
    
    return db.get_audit_log(limit, offset, action_type, user_id, start_date, end_date)

@app.get("/api/vacations")
async def get_vacations(
    request: Request,
    user_id: int = None,
    status: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Database = Depends(get_db)
):
    """Получить список отпусков"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_vacations")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    return db.get_vacations(user_id, status, start_date, end_date)

@app.post("/api/vacations")
async def create_vacation(
    request: Request,
    user_id: int,
    start_date: str,
    end_date: str,
    vacation_type: str = "annual",
    notes: str = None,
    db: Database = Depends(get_db)
):
    """Создать запись об отпуске"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="api_vacations")
    payload = authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    current_user = db.get_web_user_by_username(payload.get("sub"))
    created_by = current_user.get("id") if current_user else None
    
    vacation_id = db.create_vacation(user_id, start_date, end_date, vacation_type, created_by, notes)
    
    # Логируем действие
    db.add_audit_log_entry(
        "vacation_created",
        user_id=created_by,
        username=payload.get("sub"),
        target_type="vacation",
        target_id=vacation_id,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent")
    )
    
    return {"id": vacation_id, "status": "created"}

@app.get("/api/sick-leaves")
async def get_sick_leaves(
    request: Request,
    user_id: int = None,
    status: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Database = Depends(get_db)
):
    """Получить список больничных"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_sick_leaves")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    return db.get_sick_leaves(user_id, status, start_date, end_date)

@app.post("/api/sick-leaves")
async def create_sick_leave(
    request: Request,
    user_id: int,
    start_date: str,
    end_date: str,
    notes: str = None,
    db: Database = Depends(get_db)
):
    """Создать запись о больничном"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="api_sick_leaves")
    payload = authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    current_user = db.get_web_user_by_username(payload.get("sub"))
    created_by = current_user.get("id") if current_user else None
    
    sick_leave_id = db.create_sick_leave(user_id, start_date, end_date, created_by, notes)
    
    # Логируем действие
    db.add_audit_log_entry(
        "sick_leave_created",
        user_id=created_by,
        username=payload.get("sub"),
        target_type="sick_leave",
        target_id=sick_leave_id,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent")
    )
    
    return {"id": sick_leave_id, "status": "created"}

@app.get("/api/report-templates")
async def get_report_templates(
    request: Request,
    template_type: str = None,
    db: Database = Depends(get_db)
):
    """Получить список шаблонов отчетов"""
    rate_limit(request, max_requests=30, window_seconds=60, key_prefix="api_templates")
    authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    return db.get_report_templates(template_type)

@app.post("/api/report-templates")
async def create_report_template(
    request: Request,
    name: str,
    template_type: str,
    config: str,
    description: str = None,
    db: Database = Depends(get_db)
):
    """Создать шаблон отчета"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="api_templates")
    payload = authorize_request(request, require_roles=["admin", "manager"])
    
    current_user = db.get_web_user_by_username(payload.get("sub"))
    created_by = current_user.get("id") if current_user else None
    
    template_id = db.create_report_template(name, template_type, config, created_by, description)
    
    return {"id": template_id, "status": "created"}

@app.delete("/api/report-templates/{template_id}")
async def delete_report_template(
    request: Request,
    template_id: int,
    db: Database = Depends(get_db)
):
    """Удалить шаблон отчета"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="api_templates")
    payload = authorize_request(request, require_roles=["admin"])
    
    success = db.delete_report_template(template_id)
    if not success:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return {"status": "deleted"}

@app.post("/api/export/send-email")
async def send_report_email(
    request: Request,
    to_email: str,
    report_type: str,
    start_date: str = None,
    end_date: str = None,
    period: str = None,
    format: str = "xlsx",
    db: Database = Depends(get_db)
):
    """Отправить отчет по email"""
    rate_limit(request, max_requests=5, window_seconds=60, key_prefix="export_email")
    payload = authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    from utils.email_sender import email_sender
    from email_validator import validate_email, EmailNotValidError
    import tempfile
    
    # Валидация email
    try:
        validate_email(to_email)
    except EmailNotValidError:
        raise HTTPException(status_code=400, detail="Invalid email address")
    
    # Генерируем отчет
    today = datetime.now(timezone.utc).date()
    
    if period == "last_week":
        end = today - timedelta(days=1)
        start = end - timedelta(days=6)
    elif period == "last_month":
        end = today - timedelta(days=1)
        start = end - timedelta(days=29)
    elif period == "current_month":
        start = today.replace(day=1)
        end = today
    elif start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        raise HTTPException(status_code=400, detail="Either period or start_date+end_date must be provided")
    
    # Создаем временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{format}") as tmp_file:
        report_path = Path(tmp_file.name)
        
        if report_type == "pivot":
            report_data = db.get_pivot_report(start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
            
            if format == "xlsx":
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет"
                
                ws['A1'] = "Сотрудник"
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                ws['A1'].fill = header_fill
                ws['A1'].font = header_font
                
                col = 2
                for day in report_data['days']:
                    cell = ws.cell(row=1, column=col)
                    cell.value = datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m")
                    cell.fill = header_fill
                    cell.font = header_font
                    col += 1
                
                total_col = col
                ws.cell(row=1, column=total_col).value = "Итого"
                ws.cell(row=1, column=total_col).fill = header_fill
                ws.cell(row=1, column=total_col).font = header_font
                
                row = 2
                for employee in report_data['employees']:
                    employee_id = employee['id']
                    ws.cell(row=row, column=1).value = employee['fio']
                    
                    col = 2
                    for day in report_data['days']:
                        hours = report_data['data'][employee_id].get(day, 0)
                        ws.cell(row=row, column=col).value = format_hours_to_hhmm_util(hours)
                        col += 1
                    
                    total_hours = report_data['totals'][employee_id]
                    ws.cell(row=row, column=total_col).value = format_hours_to_hhmm_util(total_hours)
                    ws.cell(row=row, column=total_col).font = Font(bold=True)
                    row += 1
                
                ws.column_dimensions['A'].width = 25
                wb.save(report_path)
        
        # Отправляем email
        report_name = f"Отчет_{report_type}_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}"
        subject = f"Отчет: {report_name}"
        
        success = email_sender.send_report_email(
            to_email=to_email,
            subject=subject,
            report_file=report_path,
            report_name=report_name
        )
        
        # Удаляем временный файл
        try:
            report_path.unlink()
        except:
            pass
        
        if success:
            # Логируем отправку
            username = payload.get("sub", "unknown")
            log_data_export("email_report", username, {"to": to_email, "report_type": report_type})
            return {"status": "sent", "to": to_email}
        else:
            raise HTTPException(status_code=500, detail="Failed to send email")

@app.get("/api/export/ical")
async def export_ical(
    request: Request,
    user_id: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Database = Depends(get_db)
):
    """Экспорт событий в формат iCal (.ics)"""
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="export_ical")
    payload = authorize_request(request, require_roles=["admin", "manager", "hr", "user"])
    
    from icalendar import Calendar, Event
    from datetime import datetime as dt
    
    # Если user_id не указан и роль user, используем текущего пользователя
    if not user_id and payload.get("role") == "user":
        # Нужно получить user_id из токена или сессии
        # Пока используем переданный параметр
        pass
    
    if not start_date:
        start_date = (datetime.now(timezone.utc).date() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).date().strftime("%Y-%m-%d")
    
    # Получаем события
    if user_id:
        person = db.get_person_by_id(user_id)
        if not person:
            raise HTTPException(status_code=404, detail="User not found")
        events = db.get_events_by_period(person['tg_user_id'], f"{start_date}T00:00:00", f"{end_date}T23:59:59")
    else:
        # Все события за период
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT e.*, p.fio
                FROM events e
                JOIN people p ON e.user_id = p.tg_user_id
                WHERE e.ts >= ? AND e.ts <= ?
                ORDER BY e.ts
            """, (f"{start_date}T00:00:00", f"{end_date}T23:59:59"))
            events = [dict(row) for row in cursor.fetchall()]
    
    # Создаем календарь
    cal = Calendar()
    cal.add('prodid', '-//Attendance System//EN')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    
    # Группируем события по дням для создания событий работы
    daily_events = defaultdict(list)
    for event in events:
        event_date = event['ts'][:10]
        daily_events[event_date].append(event)
    
    for date_str, day_events in daily_events.items():
        checkins = [e for e in day_events if e['action'] == 'in']
        checkouts = [e for e in day_events if e['action'] == 'out']
        
        if checkins and checkouts:
            checkin_time = datetime.fromisoformat(checkins[0]['ts'].replace('Z', '+00:00'))
            checkout_time = datetime.fromisoformat(checkouts[-1]['ts'].replace('Z', '+00:00'))
            
            work_hours = (checkout_time - checkin_time).total_seconds() / 3600
            
            event = Event()
            event.add('summary', f'Рабочий день: {format_hours_to_hhmm_util(work_hours)}')
            event.add('dtstart', checkin_time)
            event.add('dtend', checkout_time)
            event.add('description', f'Приход: {checkin_time.strftime("%H:%M")}, Уход: {checkout_time.strftime("%H:%M")}')
            cal.add_component(event)
    
    # Возвращаем файл
    filename = f"attendance_{start_date}_{end_date}.ics"
    return Response(
        content=cal.to_ical(),
        media_type="text/calendar",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get(
    "/api/metrics",
    response_model=MetricsResponse,
    responses={
        200: {"description": "Success", "model": MetricsResponse},
        401: {"description": "Unauthorized", "model": ErrorResponse}
    },
    tags=["health"],
    summary="Получить метрики производительности",
    description="""
    Возвращает детальные метрики производительности системы.
    
    **Включает:**
    - Метрики базы данных (статус, размер)
    - Метрики Redis (подключение, использование памяти)
    - Системные метрики (CPU, память, диск)
    - Статистика приложения (пользователи, события, токены)
    
    **Использование:**
    - Интеграция с Prometheus/Grafana
    - Мониторинг производительности
    - Анализ использования ресурсов
    
    **Пример:**
    ```bash
    curl -H "Authorization: Bearer <token>" https://api.example.com/api/metrics
    ```
    """
)
async def get_metrics(db: Database = Depends(get_db)):
    """Get detailed performance metrics (for monitoring systems)"""
    from utils.metrics import get_system_metrics, get_redis_metrics, get_database_metrics
    
    metrics = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "database": get_database_metrics(db),
        "redis": get_redis_metrics(),
        "system": get_system_metrics()
    }
    
    # Add system health stats if available
    try:
        health_stats = db.get_system_health_stats()
        metrics["application"] = health_stats
    except:
        pass
    
    return metrics

def format_hours_to_hhmm_util(hours: float) -> str:
    """Утилита для форматирования часов в ЧЧ:ММ"""
    if hours <= 0:
        return "-"
    whole_hours = int(hours)
    minutes = int((hours - whole_hours) * 60)
    return f"{whole_hours}:{minutes:02d}"

@app.get("/api/export/pivot")
async def export_pivot_report(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    period: str = None,  # "last_week", "last_month", "current_month"
    format: str = "csv",  # "csv" or "xlsx"
    db: Database = Depends(get_db)
):
    """
    Экспорт отчета в формате сводной таблицы (pivot table).
    
    Формат: фамилии слева, дни сверху, часы на пересечении.
    
    Параметры:
    - start_date, end_date: даты в формате YYYY-MM-DD
    - period: "last_week", "last_month", "current_month" (альтернатива датам)
    - format: "csv" или "xlsx"
    """
    rate_limit(request, max_requests=10, window_seconds=60, key_prefix="export")
    payload = authorize_request(request, require_roles=["admin", "manager", "hr"])
    
    # Логируем экспорт
    username = payload.get("sub", "unknown")
    log_data_export("pivot_report", username, {"period": period, "format": format})
    
    # Определяем период
    today = datetime.now(timezone.utc).date()
    
    if period == "last_week":
        end = today - timedelta(days=1)
        start = end - timedelta(days=6)
    elif period == "last_month":
        end = today - timedelta(days=1)
        start = end - timedelta(days=29)  # Последние 30 дней
    elif period == "current_month":
        start = today.replace(day=1)
        end = today
    elif start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        raise HTTPException(status_code=400, detail="Either period or start_date+end_date must be provided")
    
    # Получаем данные
    report_data = db.get_pivot_report(start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
    
    if format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            import io
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Заголовок
            ws['A1'] = "Сотрудник"
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            ws['A1'].fill = header_fill
            ws['A1'].font = header_font
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Дни в заголовках
            col = 2
            for day in report_data['days']:
                cell = ws.cell(row=1, column=col)
                cell.value = datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col += 1
            
            # Итого
            total_col = col
            ws.cell(row=1, column=total_col).value = "Итого"
            ws.cell(row=1, column=total_col).fill = header_fill
            ws.cell(row=1, column=total_col).font = header_font
            ws.cell(row=1, column=total_col).alignment = Alignment(horizontal="center", vertical="center")
            
            # Данные
            row = 2
            for employee in report_data['employees']:
                employee_id = employee['id']
                fio = employee['fio']
                
                # Фамилия
                ws.cell(row=row, column=1).value = fio
                
                # Часы по дням
                col = 2
                for day in report_data['days']:
                    hours = report_data['data'][employee_id].get(day, 0)
                    cell = ws.cell(row=row, column=col)
                    cell.value = format_hours_to_hhmm_util(hours)
                    cell.alignment = Alignment(horizontal="center")
                    col += 1
                
                # Итого
                total_hours = report_data['totals'][employee_id]
                ws.cell(row=row, column=total_col).value = format_hours_to_hhmm_util(total_hours)
                ws.cell(row=row, column=total_col).font = Font(bold=True)
                ws.cell(row=row, column=total_col).alignment = Alignment(horizontal="center")
                
                row += 1
            
            # Автоподбор ширины колонок
            ws.column_dimensions['A'].width = 25
            for col_idx in range(2, total_col + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 12
            
            # Сохраняем в память
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"pivot_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export requires openpyxl library")
    
    else:  # CSV
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        headers = ["Сотрудник"] + [datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m") for day in report_data['days']] + ["Итого"]
        writer.writerow(headers)
        
        # Данные
        for employee in report_data['employees']:
            employee_id = employee['id']
            fio = employee['fio']
            row = [fio]
            
            for day in report_data['days']:
                hours = report_data['data'][employee_id].get(day, 0)
                row.append(format_hours_to_hhmm_util(hours))
            
            total_hours = report_data['totals'][employee_id]
            row.append(format_hours_to_hhmm_util(total_hours))
            writer.writerow(row)
        
        filename = f"pivot_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.csv"
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

@app.get("/robots.txt", response_class=PlainTextResponse)
async def robots():
    return "User-agent: *\nDisallow:\n"

@app.get("/favicon.ico")
async def favicon():
    return Response(status_code=204)

if __name__ == "__main__":
    uvicorn.run(app, host=API_HOST, port=API_PORT)
