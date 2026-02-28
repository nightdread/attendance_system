"""
Pivot report export: XLSX, CSV, MD formats.
"""
import csv
import io
from datetime import datetime

from fastapi import Response, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _header_cols(days: list) -> list:
    """Common header columns for pivot report."""
    day_headers = [datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m") for day in days]
    return (
        ["Сотрудник"]
        + day_headers
        + [
            "Итого",
            "Уходы QR",
            "Уходы по напоминанию (офис)",
            "Уходы через бота",
            "Уходы по напоминанию (удалёнка)",
            "Часы по QR",
            "Часы по напоминанию (офис)",
            "Часы через бота",
            "Часы по напоминанию (удалёнка)",
        ]
    )


def build_pivot_xlsx(
    report_data: dict,
    source_summary: dict,
    hours_summary: dict,
    start,
    end,
    format_hours,
) -> Response:
    """Build pivot report as XLSX and return Response."""
    try:
        content = _build_pivot_xlsx_bytes(
            report_data, source_summary, hours_summary, start, end, format_hours
        )
        filename = f"pivot_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl library")


def _build_pivot_xlsx_bytes(
    report_data: dict,
    source_summary: dict,
    hours_summary: dict,
    start,
    end,
    format_hours,
) -> bytes:
    """Build pivot XLSX and return as bytes. Shared by Response and file save."""
    # Same logic as build_pivot_xlsx but returns bytes
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "Сотрудник"
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for day in report_data["days"]:
        cell = ws.cell(row=1, column=col)
        cell.value = datetime.strptime(day, "%Y-%m-%d").strftime("%d.%m")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col += 1

    total_col = col
    extra_headers = [
        ("Итого", total_col),
        ("Уходы QR", total_col + 1),
        ("Уходы по напоминанию (офис)", total_col + 2),
        ("Уходы через бота", total_col + 3),
        ("Уходы по напоминанию (удалёнка)", total_col + 4),
        ("Часы по QR", total_col + 5),
        ("Часы по напоминанию (офис)", total_col + 6),
        ("Часы через бота", total_col + 7),
        ("Часы по напоминанию (удалёнка)", total_col + 8),
    ]
    hours_qr_col = total_col + 5
    hours_reminder_office_col = total_col + 6
    hours_bot_remote_col = total_col + 7
    hours_reminder_remote_col = total_col + 8

    for title, c in extra_headers:
        ws.cell(row=1, column=c).value = title
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for employee in report_data["employees"]:
        employee_id = employee["id"]
        fio = employee["fio"]

        ws.cell(row=row, column=1).value = fio

        col = 2
        for day in report_data["days"]:
            hours = report_data["data"][employee_id].get(day, 0)
            cell = ws.cell(row=row, column=col)
            cell.value = format_hours(hours)
            cell.alignment = Alignment(horizontal="center")
            col += 1

        total_hours = report_data["totals"][employee_id]
        ws.cell(row=row, column=total_col).value = format_hours(total_hours)
        ws.cell(row=row, column=total_col).font = Font(bold=True)
        ws.cell(row=row, column=total_col).alignment = Alignment(horizontal="center")

        source_row = source_summary.get(employee_id, {})
        hours_row = hours_summary.get(employee_id, {})
        ws.cell(row=row, column=total_col + 1).value = source_row.get("checkout_qr", 0)
        ws.cell(row=row, column=total_col + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=total_col + 2).value = source_row.get("checkout_reminder_office", 0)
        ws.cell(row=row, column=total_col + 2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=total_col + 3).value = source_row.get("checkout_bot_remote", 0)
        ws.cell(row=row, column=total_col + 3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=total_col + 4).value = source_row.get("checkout_reminder_remote", 0)
        ws.cell(row=row, column=total_col + 4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=hours_qr_col).value = format_hours(hours_row.get("hours_qr", 0))
        ws.cell(row=row, column=hours_qr_col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=hours_reminder_office_col).value = format_hours(
            hours_row.get("hours_reminder_office", 0)
        )
        ws.cell(row=row, column=hours_reminder_office_col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=hours_bot_remote_col).value = format_hours(hours_row.get("hours_bot_remote", 0))
        ws.cell(row=row, column=hours_bot_remote_col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=hours_reminder_remote_col).value = format_hours(
            hours_row.get("hours_reminder_remote", 0)
        )
        ws.cell(row=row, column=hours_reminder_remote_col).alignment = Alignment(horizontal="center")

        row += 1

    ws.column_dimensions["A"].width = 25
    for col_idx in range(2, hours_reminder_remote_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def save_pivot_xlsx_to_path(
    report_data: dict,
    source_summary: dict,
    hours_summary: dict,
    start,
    end,
    format_hours,
    path,
) -> None:
    """Build pivot report as XLSX and save to file path. For email/save use cases."""
    content = _build_pivot_xlsx_bytes(
        report_data, source_summary, hours_summary, start, end, format_hours
    )
    with open(path, "wb") as f:
        f.write(content)


def build_pivot_csv(
    report_data: dict,
    source_summary: dict,
    hours_summary: dict,
    start,
    end,
    format_hours,
) -> Response:
    """Build pivot report as CSV and return Response."""
    output = io.StringIO()
    writer = csv.writer(output)

    headers = _header_cols(report_data["days"])
    writer.writerow(headers)

    for employee in report_data["employees"]:
        employee_id = employee["id"]
        fio = employee["fio"]
        row = [fio]

        for day in report_data["days"]:
            hours = report_data["data"][employee_id].get(day, 0)
            row.append(format_hours(hours))

        total_hours = report_data["totals"][employee_id]
        source_row = source_summary.get(employee_id, {})
        hours_row = hours_summary.get(employee_id, {})

        row.append(format_hours(total_hours))
        row.append(source_row.get("checkout_qr", 0))
        row.append(source_row.get("checkout_reminder_office", 0))
        row.append(source_row.get("checkout_bot_remote", 0))
        row.append(source_row.get("checkout_reminder_remote", 0))
        row.append(format_hours(hours_row.get("hours_qr", 0)))
        row.append(format_hours(hours_row.get("hours_reminder_office", 0)))
        row.append(format_hours(hours_row.get("hours_bot_remote", 0)))
        row.append(format_hours(hours_row.get("hours_reminder_remote", 0)))
        writer.writerow(row)

    filename = f"pivot_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def build_pivot_md(
    report_data: dict,
    source_summary: dict,
    hours_summary: dict,
    start,
    end,
    format_hours,
) -> Response:
    """Build pivot report as Markdown and return Response."""
    header_cols = _header_cols(report_data["days"])
    lines = [
        f"# Отчет {start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}",
        "",
        "| " + " | ".join(header_cols) + " |",
        "|---|" + "|".join("---:" for _ in range(len(header_cols) - 1)) + "|",
    ]

    for employee in report_data["employees"]:
        employee_id = employee["id"]
        fio = employee["fio"]
        row_values = [fio]

        for day in report_data["days"]:
            hours = report_data["data"][employee_id].get(day, 0)
            row_values.append(format_hours(hours))

        total_hours = report_data["totals"][employee_id]
        source_row = source_summary.get(employee_id, {})
        hours_row = hours_summary.get(employee_id, {})
        row_values.append(format_hours(total_hours))
        row_values.append(str(source_row.get("checkout_qr", 0)))
        row_values.append(str(source_row.get("checkout_reminder_office", 0)))
        row_values.append(str(source_row.get("checkout_bot_remote", 0)))
        row_values.append(str(source_row.get("checkout_reminder_remote", 0)))
        row_values.append(format_hours(hours_row.get("hours_qr", 0)))
        row_values.append(format_hours(hours_row.get("hours_reminder_office", 0)))
        row_values.append(format_hours(hours_row.get("hours_bot_remote", 0)))
        row_values.append(format_hours(hours_row.get("hours_reminder_remote", 0)))
        lines.append("| " + " | ".join(row_values) + " |")

    content = "\n".join(lines) + "\n"
    filename = f"pivot_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.md"
    return Response(
        content=content,
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
